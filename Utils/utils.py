import openpyxl
import logging
from openpyxl.styles import Font, Color
import pandas as pd

# read data from the excel
def read_data(file_path ,sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.cell(row, column).value

# row count from the excel
def row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.max_row

# write data to the excel sheet
def write_data(file_path,sheet_name, row, column,data,col2=None, value2=None):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    work_sheet.cell(row,column).value = data
    if col2 and value2 is not None:
        work_sheet.cell(row=row, column=col2).value=value2
    return workbook.save(file_path)
