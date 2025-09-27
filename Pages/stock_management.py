from selenium.webdriver.common.by import By
from selenium.webdriver.common.log import Log
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException, InvalidElementStateException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
import traceback
import time
from datetime import datetime
import openpyxl

class StockManagement:
    def __init__(self, driver):
        self.driver = driver
        self.stock_link_xpath = (By.XPATH,"//div[contains(@class, 'sidebar-item-container') and contains(@item-public, '1')]/descendant::a[contains(@title, 'Stock')]")
        self.stock_workspace_scroll = (By.XPATH, "//div[@class='col layout-main-section-wrapper']/child::div[@class='layout-main-section']")
        self.item_click_xpath = (By.XPATH, "//div[@class='ce-block__content']/child::div[contains(@shortcut_name, 'Item')]/descendant::span[contains(@title, 'Item') and contains(@class, 'ellipsis')]")
        # self.filter_set_xpath = (By.XPATH, "//div[@class='filter-selector']/descendant::button[contains(@class, 'filter-button') and contains(@class, 'btn-primary-light')]")
        # self.add_filter_btn = (By.XPATH, "//div[contains(@class, 'popover-body') and contains(@class, 'popover-content')]/descendant::div[contains(@class, 'filter-action-buttons') and contains(@class, 'mt-2')]//button[contains(text(), '+ Add a Filter')]")

        # add item btn
        self.item_add_btn = (By.XPATH, "//button[contains(@data-label,'Add Item')]")
        self.item_add_full_form = (By.XPATH, "//div[@class='modal-dialog']//div[@class='modal-content']/child::div[@class='modal-footer']//div[@class='custom-actions']//button")
        self.add_item_popup = (By.CSS_SELECTOR, "div.modal.show div.modal-dialog")
        self.item_save_popup = (By.XPATH, "//div[@role='dialog']//button[@type='button'][normalize-space()='Save']")
        self.item_code__two_xpath = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'item_code')]")
        self.item_code_xpath = (By.XPATH,"//div[@role='dialog']//div//div//div//div//div//div//div//div//div//form//div[@data-fieldtype='Data']//div//div//div//input[@type='text']")
        self.item_group_xpath = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'item_group')]")
        self.item_group_options = (By.XPATH, "//div[@class='awesomplete']/child::ul[@id='awesomplete_list_3']//div[@role='option']//p")
        self.uom = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'stock_uom')]")
        self.is_stock_item = (By.XPATH, "//input[@class='input-with-feedback bold']")
        self.opening_stock =(By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'opening_stock')]")
        self.valuation_rate = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'valuation_rate')]")
        self.selling_rate = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'standard_rate')]")
        self.collapse_disable = (By.XPATH, "//div[@data-fieldname='section_break_11']/child::div[contains(@class,'section-head') and contains(@class, 'collapsed')]")
        self.collapse_enable = (By.XPATH, "//div[@data-fieldname='section_break_11']/child::div[contains(@class,'section-head') and contains(@class, 'collapsible')]")
        self.description = (By.XPATH, "//div[@class='ql-container ql-snow']//div[@class='ql-editor ql-blank']")
        self.inventory_tab = (By.XPATH, "//ul[@id='form-tabs']/li[contains(@class, 'nav-item') and contains(@class, 'show')]//a[@data-fieldname='inventory_section']")
        self.shelf_days = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'shelf_life_in_days')]")
        self.end_of_life = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'end_of_life')]")
        self.default_mr_type = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[contains(@class, 'control-input')and contains(@class,'flex') and contains(@class, 'align-center')]//select[contains(@data-fieldname,'default_material_request_type')]")
        self.valuation_method = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[contains(@class, 'control-input')and contains(@class,'flex') and contains(@class, 'align-center')]//select[contains(@data-fieldname,'valuation_method')]")
        self.warranty_period = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'warranty_period')]")
        self.weight_per_unit = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'weight_per_unit')]")
        self.weight_uom = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'weight_uom')]")
        self.barcode_row_add_btn = (By.XPATH, "//div[@data-fieldname='barcodes']//button[@type='button'][normalize-space()='Add Row']")
        self.child_table_edit_btn = (By.XPATH, "//div[@class='grid-body']//div[@class='rows']//div[@class='grid-row']//div[contains(@class,'data-row') and contains(@class,'row')]//div[@class='col']")
        self.barcode_popup = (By.XPATH, "//div[@class='form-in-grid']")
        self.barcode_sn = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'barcode')]")
        self.barcode_type = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[contains(@class, 'control-input')and contains(@class,'flex') and contains(@class, 'align-center')]//select[contains(@data-fieldname,'barcode_type')]")
        self.barcode_uom = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'uom')and contains(@data-doctype, 'Item Barcode')]")
        self.close_child_tbl_row = (By.XPATH, "//button[@class='btn btn-secondary btn-sm pull-right grid-collapse-row']")
        self.auto_reorder_collapse_click = (By.XPATH, "//div[@data-fieldname='reorder_section']/child::div[contains(@class,'section-head') and contains(@class, 'collapsed')]")
        self.auto_reorder_collapse_visible = (By.XPATH, "//div[@data-fieldname='reorder_section']/child::div[contains(@class,'section-head') and contains(@class, 'collapsible')]")
        self.auto_reorder_add_row = (By.XPATH, "//div[@data-fieldname='reorder_levels']//button[@type='button'][normalize-space()='Add Row']")
        self.auto_reorder_row_edit = (By.XPATH, "//div[@data-fieldname='__column_9']/descendant::div[@class='col']")
        self.reorder_checkin_warehouse = (By.XPATH,"//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'warehouse_group')]")
        self.order_warehouse_option = (By.XPATH,"//div[@class='awesomplete']/child::ul[@id='awesomplete_list_23']//div[@role='option']//p")
        self.reorder_request_for = (By.XPATH, "//div[@data-fieldname='warehouse']//input[@role='combobox']")
        self.reorder_request_for_option = (By.XPATH,"//div[@class='awesomplete']/child::ul[@id='awesomplete_list_24']//div[@role='option']//p")
        self.reorder_level = (By.XPATH, "//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'warehouse_reorder_level')]")
        self.reorder_warehouse_qty = (By.XPATH,"//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[@class='control-input']//input[contains(@data-fieldname,'warehouse_reorder_qty')]")
        self.reorder_purpose = (By.XPATH,"//div[@class='form-group']/child::div[@class='control-input-wrapper']//div[contains(@class, 'control-input')and contains(@class,'flex') and contains(@class, 'align-center')]//select[contains(@data-fieldname,'material_request_type') and contains(@data-doctype,'Item Reorder')]")
        self.auto_reorder_tbl_close = (By.XPATH, "//div[@class='grid-row grid-row-open']//button[@class='btn btn-secondary btn-sm pull-right grid-collapse-row']")
        self.barcode_error_msg = (By.XPATH, "//div[@class='modal-dialog msgprint-dialog']//div[@class='modal-content']/descendant::div[@class='msgprint']")
        self.success_msg = (By.XPATH, "//div[@id='dialog-container']//div[@id='alert-container']/descendant::div[@class='alert-message']")


        # self.save_btn_xpath = (By.XPATH, "//div[@class='container']/child::div[contains(@class, 'row') and contains(@class, 'flex') and contains(@class, 'align-center') and contains(@class, 'justify-between') and contains(@class, 'page-head-content')]/child::div[contains(@class, 'col') and contains(@class, 'flex') and contains(@class, 'justify-content-end') and contains(@class, 'page-actions')]/descendant::button[contains(@data-label, 'Save')]")
        self.save_btn_xpath = (By.XPATH, "//div[@id='page-Item']//button[@data-label='Save']")

        self.check_item_enable = (By.XPATH, "//span[@class='indicator-pill no-indicator-dot whitespace-nowrap blue']//span[contains(text(),'Enabled')]")
    
    def stock_page_access(self):
        WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(self.stock_link_xpath))
        side_menu_link = self.driver.find_element(*self.stock_link_xpath)
        side_menu_link.click()
        time.sleep(3)
        workspace_scroll = self.driver.find_element(*self.stock_workspace_scroll)
        scroll_origin = ScrollOrigin.from_element(workspace_scroll)
        ActionChains(self.driver).scroll_from_origin(scroll_origin,0, 100).perform()

        WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(self.item_click_xpath))
        move_to_ele = self.driver.find_element(*self.item_click_xpath)
        ActionChains(self.driver).move_to_element(move_to_ele).perform()
        self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/item_text.png")
        time.sleep(3)
        move_to_ele.click()
        time.sleep(2)

    def add_item(self, excel_sheet_path, sheet_name, row):
        add_item = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.item_add_btn))
        self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/add_item_btn_click.png")
        add_item.click()
        
        # check popup open
        # popup = WebDriverWait(self.driver, 20).until(
        #     EC.visibility_of_element_located(self.add_item_popup)
        # )
        # self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/item_popup.png")
        # popup.find_element(*self.item_code_xpath).send_keys("Sony Earbuds")
        # print("Popup Element is Visible..")

        # time.sleep(5)
        # save_btn = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.item_save_popup))
        # save_btn.click()

        full_form_open = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.item_add_full_form))
        full_form_open.click()
        item_datas = get_item_master_datas(excel_sheet_path, sheet_name, row)
        print("Item Datas", item_datas)
        for item in item_datas:
            try:
                if item['code']:
                    item_code = self.driver.find_element(*self.item_code__two_xpath)
                    item_code.send_keys(item['code'])
                # item_group = self.driver.find_element(*self.item_group_xpath)
                item_group = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.item_group_xpath))
                item_group.clear()
                item_group.send_keys(item['group'])
                item_group.send_keys(Keys.ARROW_DOWN)
                time.sleep(0.5)
                item_group.send_keys(Keys.ENTER)
                time.sleep(2)

                uom = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable(self.uom))
                uom.clear()
                uom.send_keys(item['uom'])
                uom.send_keys(Keys.ARROW_DOWN)
                time.sleep(0.5)
                uom.send_keys(Keys.ENTER)
                time.sleep(2)

                # maintain_stock = self.driver.find_element(self.is_stock_item)
                maintain_stock = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.is_stock_item))
                maintain_stock.click()

                if item['opening_stock']:
                    opening_stock = self.driver.find_element(*self.opening_stock)
                    opening_stock.send_keys(item['opening_stock'])

                
                if item['valuation_rate']:
                    value_rate = self.driver.find_element(*self.valuation_rate)
                    value_rate.send_keys(item['valuation_rate'])

                if item['selling_rate']:
                    selling_rate = self.driver.find_element(*self.selling_rate)
                    selling_rate.send_keys(item['selling_rate'])

                open_desc = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.collapse_disable))
                open_desc.click()
                time.sleep(2)
                open_enable = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.collapse_enable))
                self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/description.png")
                time.sleep(2)
                description = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.description))
                description.clear()
                description.click()
                ActionChains(self.driver).move_to_element(description)
                description.send_keys(item['description'])
                time.sleep(2)

                # inventory tab
                try:
                    inv_tab = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(self.inventory_tab))
                    inv_tab.click()
                    time.sleep(0.5)
                    # self.driver.implicitly_wait(2)
                    self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/inventory_tab.png")

                    # shelf in days
                    shelf_in_days = WebDriverWait(self.driver,5).until(EC.presence_of_element_located(self.shelf_days))
                    shelf_in_days.clear()
                    shelf_in_days.send_keys(item['shelf_in_days'])

                    end_life = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable(self.end_of_life))
                    end_life.click()
                    end_life.clear()
                    date_value = item['end_of_life']
                    if isinstance(date_value, datetime):
                        date_val = date_value.strftime("%d-%m-%Y")
                    end_life.send_keys(date_val)

                    default_mrq_type = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.default_mr_type))
                    select_mr_type = Select(default_mrq_type)
                    select_mr_type.select_by_value(item['default_mr_type'])

                    value_method = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.valuation_method))
                    select_value_method = Select(value_method)
                    select_value_method.select_by_value(item['valuation_method'])

                    warranty_days = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.warranty_period))
                    warranty_days.clear()
                    warranty_days.send_keys(item['warranty_days'])

                    weight_per = WebDriverWait(self.driver,5).until(EC.presence_of_element_located(self.weight_per_unit))
                    weight_per.clear()
                    weight_per.send_keys(item['weight_per_unit'])

                    weight_uom = WebDriverWait(self.driver,5).until(EC.presence_of_element_located(self.weight_uom))
                    weight_uom.clear()
                    weight_uom.send_keys(item['weight_uom'])
                    weight_uom.send_keys(Keys.ENTER)

                    if item['barcode_sn'] and item['barcode_type'] and item['barcode_uom']:
                        barcode_add = WebDriverWait(self.driver,5).until(EC.element_to_be_clickable(self.barcode_row_add_btn))
                        barcode_add.click()
                        self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/barcode_add_empty.png")

                        WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(self.barcode_popup))
                        time.sleep(0.4)

                        barcode_num = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.barcode_sn))
                        barcode_num.clear()
                        barcode_num.send_keys(item['barcode_sn'])

                        barcode_sn_type = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.barcode_type))
                        select_barcode_type = Select(barcode_sn_type)
                        select_barcode_type.select_by_value(item['barcode_type'])

                        barcode_uoms = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.barcode_uom))
                        barcode_uoms.clear()
                        barcode_uoms.send_keys(item['barcode_uom'])
                        barcode_uoms.send_keys(Keys.ENTER)

                        close_child_barcode= WebDriverWait(self.driver,5).until(EC.element_to_be_clickable(self.close_child_tbl_row))
                        close_child_barcode.click()

                    open_auto_reorder = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.auto_reorder_collapse_click))
                    open_auto_reorder.click()
                    time.sleep(2)
                    WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.auto_reorder_collapse_visible))
                    self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/auto_reorder.png")
                    time.sleep(2)

                    add_auto_reorder = WebDriverWait(self.driver,5).until(EC.element_to_be_clickable(self.auto_reorder_add_row))
                    add_auto_reorder.click()
                    edit_row = WebDriverWait(self.driver,10).until(EC.presence_of_element_located(self.auto_reorder_row_edit))
                    edit_row.click()
                    self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/auto_reorder_edit_row.png")

                    check_in_group = WebDriverWait(self.driver,10).until(EC.presence_of_element_located(self.reorder_checkin_warehouse))
                    check_in_group.clear()
                    check_in_group.send_keys(item['checkin_warehouse'])
                    check_in_group.send_keys(Keys.ARROW_DOWN)
                    time.sleep(0.5)
                    check_in_group.send_keys(Keys.ENTER)
                    # time.sleep(2)

                    request_for_wareshouse = WebDriverWait(self.driver,10).until(EC.presence_of_element_located(self.reorder_request_for))
                    request_for_wareshouse.clear()
                    request_for_wareshouse.send_keys(item['request_for_wh'])
                    request_for_wareshouse.send_keys(Keys.ARROW_DOWN)
                    time.sleep(0.3)
                    request_for_wareshouse.send_keys(Keys.ENTER)
                    time.sleep(2)

                    re_order_level_reach = WebDriverWait(self.driver,10).until(EC.presence_of_element_located(self.reorder_level))
                    re_order_level_reach.clear()
                    re_order_level_reach.send_keys(item['reorder_level'])

                    re_order_qty = WebDriverWait(self.driver,5).until(EC.presence_of_element_located(self.reorder_warehouse_qty))
                    re_order_qty.clear()
                    re_order_qty.send_keys(item['reorder_qty'])

                    material_type = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.reorder_purpose))
                    material_type_select = Select(material_type)
                    material_type_select.select_by_value(item['mr_type'])

                    close_auto_reoder_table = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.auto_reorder_tbl_close))
                    close_auto_reoder_table.click()
                    time.sleep(2)
                except Exception as e:

                    ActionChains(self.driver).scroll_by_amount(0,-50).perform()
                    time.sleep(4)
                    save_btn = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.save_btn_xpath))
                    save_btn.click()
                    time.sleep(0.5)

                error_msg = None
                item_saved_msg = None
                self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/barcode_error.png")
                try:
                    barcode_error = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(self.barcode_error_msg))
                    error_msg = barcode_error.text
                    print("Error Msg =>", error_msg)
                    time.sleep(5)
                except Exception as e:
                    pass

                try:
                    enable_text=WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(self.check_item_enable))
                    # alert_found = self.driver.switch_to.alert
                    success = enable_text.text
                    print("Saved =>", success)
                    time.sleep(5)
                    self.driver.get_screenshot_as_file("E:/Erpnext Automation/Screenshots/item_added.png")
                except Exception as e:
                    pass
                return {"error": error_msg, "success": success}
            except Exception as e:
                print("Exception Type", type(e))
                print("Exception Name:", e.__class__.__name__)
                tb = traceback.extract_tb(e.__traceback__)
                for frame in tb:
                    print(f"File: {frame.filename}, Line: {frame.lineno}, Code: {frame.line}")
        time.sleep(5)

def get_item_master_datas(excel_file_path, sheet_name, start_row):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]

    item_data = []
    for i in range(start_row, sheet.max_row+1):
        item_ = {
            "code" : sheet.cell(row=i, column=5).value,
            "item_name" : sheet.cell(row=i, column=6).value,
            "uom" : sheet.cell(row=i, column=7).value,
            "description" : sheet.cell(row=i, column=8).value,
            "group" : sheet.cell(row=i, column=9).value,
            "opening_stock" : sheet.cell(row=i, column=10).value,
            "valuation_rate" : sheet.cell(row=i, column=11).value,
            "selling_rate" : sheet.cell(row=i, column=12).value,
            "shelf_in_days":sheet.cell(row=i, column=13).value,
            "end_of_life":sheet.cell(row=i, column=14).value,
            "default_mr_type":sheet.cell(row=i, column=15).value,
            "valuation_method":sheet.cell(row=i, column=16).value,
            "warranty_days":sheet.cell(row=i, column=17).value,
            "weight_per_unit":sheet.cell(row=i, column=18).value,
            "weight_uom":sheet.cell(row=i, column=19).value,
            "barcode_sn":sheet.cell(row=1, column=20).value,
            "barcode_type":sheet.cell(row=i, column=21).value,
            "barcode_uom":sheet.cell(row=i, column=22).value,
            "checkin_warehouse":sheet.cell(row=i, column=23).value,
            "request_for_wh":sheet.cell(row=i, column=24).value,
            "reorder_level":sheet.cell(row=i, column=25).value,
            "reorder_qty":sheet.cell(row=i, column=26).value,
            "mr_type":sheet.cell(row=i, column=27).value
        }

        item_data.append(item_)
    return item_data